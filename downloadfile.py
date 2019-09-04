import openpyxl


def createExcel(phoneList):
    # 建立Excel
    wb = openpyxl.Workbook()
    # 新建sheet表
    wb.create_sheet(index=0, title="userData")  # 可通过index控制创建的表的位置
    # 获取所有表名
    sheet_names = wb.sheetnames
    # 根据表名打开sheet表
    sheet1 = wb[sheet_names[0]]  # 打开第一个 sheet 工作表

    rowNum = 2
    sheet1.cell(row=1, column=1, value="手机号")
    for phone in phoneList:
        sheet1.cell(row=rowNum, column=1, value=phone)
        rowNum += 1
    wb.save(r'/home/admin/deploy/python_src/kouzi_web/static/upload/users.xls')
